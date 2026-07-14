from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_review_store import AiReviewAuditRecord
from .models import OperationRecord
from .scanner import STATUS_FOLDERS, scan_all_counts

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")


def _style_sheet(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_report(
    root: Path | str,
    output_path: Path | str,
    session_records: list[OperationRecord],
    ai_reviews: list[AiReviewAuditRecord] | None = None,
) -> Path:
    root_path = Path(root)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    ai_reviews = list(ai_reviews or [])

    counts = scan_all_counts(root_path)
    failed_records = [record for record in session_records if record.action == "不通过"]
    issues_by_person: dict[str, Counter[str]] = defaultdict(Counter)
    all_issues: Counter[str] = Counter()
    for record in failed_records:
        for issue in record.issues or ["其他问题"]:
            issues_by_person[record.person][issue] += 1
            all_issues[issue] += 1

    ai_by_person: dict[str, list[AiReviewAuditRecord]] = defaultdict(list)
    ai_issues_by_person: dict[str, Counter[str]] = defaultdict(Counter)
    all_ai_issues: Counter[str] = Counter()
    for record in ai_reviews:
        ai_by_person[record.person].append(record)
        for issue in record.result.issue_categories:
            ai_issues_by_person[record.person][issue] += 1
            all_ai_issues[issue] += 1

    wb = Workbook()
    summary = wb.active
    summary.title = "人员汇总"
    summary_headers = [
        "人员姓名",
        "标注总数量",
        "质检完成数量",
        "待返修数量",
        "待质检数量",
        "返修提交数量",
        "质检不通过的主要问题",
        "AI检测数量",
        "AI建议通过",
        "AI建议复核",
        "AI建议返修",
        "AI高频问题",
    ]
    summary.append(summary_headers)

    all_people = sorted(
        set(counts) | set(issues_by_person) | set(ai_by_person), key=str.lower
    )
    for person in all_people:
        person_counts = counts.get(person, {status: 0 for status in STATUS_FOLDERS})
        total = sum(person_counts.get(status, 0) for status in STATUS_FOLDERS)
        issue_summary = "；".join(
            f"{name} {number}" for name, number in issues_by_person[person].most_common()
        )
        person_ai = ai_by_person.get(person, [])
        recommendations = Counter(item.result.recommendation for item in person_ai)
        ai_issue_summary = "；".join(
            f"{name} {number}" for name, number in ai_issues_by_person[person].most_common()
        )
        summary.append([
            person,
            total,
            person_counts.get("质检完成", 0),
            person_counts.get("待返修", 0),
            person_counts.get("待质检", 0),
            person_counts.get("返修提交", 0),
            issue_summary,
            len(person_ai),
            recommendations.get("pass", 0),
            recommendations.get("review", 0),
            recommendations.get("repair", 0),
            ai_issue_summary,
        ])
    _style_sheet(summary, [16, 14, 16, 14, 14, 16, 42, 14, 14, 14, 14, 42])

    detail = wb.create_sheet("不通过明细")
    detail.append([
        "质检时间",
        "人员姓名",
        "数据编号",
        "原始状态",
        "主要问题",
        "详细返修备注",
        "AI评分",
        "AI风险",
        "AI建议",
        "AI来源",
        "AI问题标签",
        "AI总结",
    ])
    for record in failed_records:
        detail.append([
            record.timestamp,
            record.person,
            record.group_name,
            record.source_status,
            "、".join(record.issues) if record.issues else "其他问题",
            record.remark,
            record.ai_score,
            _risk_text(record.ai_risk),
            _recommendation_text(record.ai_recommendation),
            _provider_text(record.ai_provider),
            "、".join(record.ai_issues),
            record.ai_summary,
        ])
    _style_sheet(detail, [22, 16, 18, 14, 34, 60, 12, 12, 14, 16, 34, 50])

    total_sheet = wb.create_sheet("总体汇总")
    grand = {
        status: sum(person.get(status, 0) for person in counts.values())
        for status in STATUS_FOLDERS
    }
    total_count = sum(grand.values())
    total_sheet.append(["统计项目", "数量"])
    total_sheet.append(["全部标注数量", total_count])
    for status in STATUS_FOLDERS:
        total_sheet.append([status, grand[status]])
    total_sheet.append(["本次质检不通过操作数", len(failed_records)])
    total_sheet.append(["AI检测数据组数", len(ai_reviews)])
    total_sheet.append(["AI建议通过", sum(item.result.recommendation == "pass" for item in ai_reviews)])
    total_sheet.append(["AI建议复核", sum(item.result.recommendation == "review" for item in ai_reviews)])
    total_sheet.append(["AI建议返修", sum(item.result.recommendation == "repair" for item in ai_reviews)])
    total_sheet.append([])
    total_sheet.append(["人工质检主要问题", "出现次数"])
    for issue, number in all_issues.most_common():
        total_sheet.append([issue, number])
    for cell in total_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_number in range(2, total_sheet.max_row + 1):
        if total_sheet.cell(row_number, 1).value == "人工质检主要问题":
            for cell in total_sheet[row_number]:
                cell.fill = SUB_FILL
                cell.font = Font(bold=True)
    total_sheet.column_dimensions["A"].width = 30
    total_sheet.column_dimensions["B"].width = 16
    for row in total_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ai_sheet = wb.create_sheet("AI辅助统计")
    ai_sheet.append([
        "检测时间",
        "人员姓名",
        "数据编号",
        "检测阶段",
        "提供商",
        "一致性评分",
        "风险等级",
        "建议结论",
        "问题标签",
        "综合说明",
    ])
    for record in ai_reviews:
        ai_sheet.append([
            record.timestamp,
            record.person,
            record.group_name,
            "在线深度复核" if record.result.stage == "online" else "本地初筛",
            _provider_text(record.result.provider),
            record.result.score,
            _risk_text(record.result.risk),
            _recommendation_text(record.result.recommendation),
            "、".join(record.result.issue_categories),
            record.result.summary,
        ])
    _style_sheet(ai_sheet, [22, 16, 18, 16, 16, 14, 14, 14, 40, 60])
    summary_row = ai_sheet.max_row + 2
    ai_sheet.cell(summary_row, 1, "AI问题分类汇总")
    ai_sheet.cell(summary_row, 2, "出现次数")
    for cell in ai_sheet[summary_row]:
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
    if all_ai_issues:
        for issue, number in all_ai_issues.most_common():
            ai_sheet.append([issue, number])
    else:
        ai_sheet.append(["无AI问题记录", 0])

    wb.save(output)
    return output


def _risk_text(value: str) -> str:
    return {"low": "低风险", "medium": "中风险", "high": "高风险"}.get(value, value)


def _recommendation_text(value: str) -> str:
    return {"pass": "建议通过", "review": "建议复核", "repair": "建议返修"}.get(value, value)


def _provider_text(value: str) -> str:
    return {
        "local": "本地规则",
        "openai": "OpenAI",
        "gemini": "Gemini",
        "custom": "自定义API",
    }.get(value, value)
