from pathlib import Path

from openpyxl import load_workbook

from app.models import OperationRecord
from app.reports import export_report


def make_group(root: Path, status: str, person: str, group: str) -> None:
    folder = root / status / person / group
    folder.mkdir(parents=True)
    (folder / f"{group}.jpg").write_bytes(b"jpg")


def test_export_report_contains_counts_totals_and_failure_details(tmp_path: Path):
    make_group(tmp_path, "质检完成", "张三", "000001")
    make_group(tmp_path, "待返修", "张三", "000002")
    make_group(tmp_path, "待质检", "李四", "000003")
    output = tmp_path / "report.xlsx"
    records = [
        OperationRecord(
            timestamp="2026-07-14 20:00:00",
            action="不通过",
            person="张三",
            group_name="000002",
            source_status="待质检",
            source_path="x",
            destination_path="y",
            issues=["细节丢失", "主体不一致"],
            remark="标签缺失",
        )
    ]

    export_report(tmp_path, output, records)

    wb = load_workbook(output)
    assert wb.sheetnames == ["人员汇总", "不通过明细", "总体汇总", "AI辅助统计"]
    summary = wb["人员汇总"]
    rows = list(summary.iter_rows(min_row=2, values_only=True))
    by_person = {row[0]: row for row in rows}
    assert by_person["张三"][1] == 2
    assert by_person["张三"][2:6] == (1, 1, 0, 0)
    assert "细节丢失 1" in by_person["张三"][6]
    assert wb["不通过明细"]["F2"].value == "标签缺失"
    assert wb["总体汇总"]["B2"].value == 3


def test_export_report_contains_ai_review_statistics(tmp_path: Path):
    from app.ai_review_models import AiReviewResult, ReviewFinding
    from app.ai_review_store import AiReviewAuditRecord

    make_group(tmp_path, "待质检", "张三", "000010")
    output = tmp_path / "ai-report.xlsx"
    ai_reviews = [
        AiReviewAuditRecord(
            timestamp="2026-07-15 10:30:00",
            status="待质检",
            person="张三",
            group_name="000010",
            folder="x",
            signature="sig",
            result=AiReviewResult(
                stage="online",
                provider="openai",
                score=48.5,
                risk="high",
                recommendation="repair",
                summary="结构与Logo异常",
                findings=[
                    ReviewFinding("结构变形", "high", 96, "右侧畸形"),
                    ReviewFinding("文字或Logo错误", "high", 91, "Logo拼写错误"),
                ],
            ),
        )
    ]

    export_report(tmp_path, output, [], ai_reviews=ai_reviews)

    wb = load_workbook(output)
    summary = wb["人员汇总"]
    headers = [cell.value for cell in summary[1]]
    assert "AI检测数量" in headers
    row = next(row for row in summary.iter_rows(min_row=2, values_only=True) if row[0] == "张三")
    assert row[headers.index("AI检测数量")] == 1
    assert row[headers.index("AI建议返修")] == 1
    ai_sheet = wb["AI辅助统计"]
    assert ai_sheet["A2"].value == "2026-07-15 10:30:00"
    assert ai_sheet["F2"].value == 48.5
    assert "结构变形" in ai_sheet["I2"].value
    values = [cell.value for cell in ai_sheet[ai_sheet.max_row]]
    assert "文字或Logo错误" in values
